import requests
import re
import openpyxl
import time

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36',
    'Referer': 'https://missav.ws/',
    'Cookie': 'dom3ic8zudi28v8lr6fgphwffqoz0j6c=eff981d2-4b16-470c-8657-6c24b290b988:3:1; localized=1; _ga=GA1.2.7467935.1670460630; _ga_Z3V6T9VBM6=GS1.1.1670512561.5.1.1670516155.0.0.0; __cf_bm=UtKX8sOkgOMJinKc8VXqsvqA5IcYcPiqbsNUPAkPaTw-1670628610-0-AT0f7fWU8xSd4YhWjCTO7HtRM/fWFVtYWvUmM/AcWi64rnENHRCxngKhXJYDGqAsAU4PBHo9sNCpXn2CvJ3r5UrMiVFMTm+5sLIjO5AKcjGsDh88FpQ69rZOt542TeVY+CtYEPUGMIGMm4lFONC7DK4=; search_history=["%E6%9D%BE%E6%9C%AC%E3%81%84%E3%81%A1%E3%81%8B","DVDES-644","%E4%BD%93%E6%A3%80","%E6%80%A7%E6%95%99%E8%82%B2","MTALL-031"]; XSRF-TOKEN=eyJpdiI6IisrZU5CUHlSOW8xaEhRUHRSdUhIMnc9PSIsInZhbHVlIjoiQTNVM1FYbWRkRWlaaWlibTNRclFOZGx5UDVyallENml5VlYxTy9ubmF5QlJLcHBBZFM2YkZ5VWQzWVFaYUZhaW94dDJkSVpIL2Rvd3FzbU9aS2YzODZyRHVJSTNrVytBWFd0TFVSV1NzZlNTME05YWFoSnBNRDcyalVva1dLaFgiLCJtYWMiOiIyOWJlZTI5NDBiYTkxZjk4OTU1M2U3YzAyNjg0ZThiZmFjMmZkMzc3YTY0NzU3ZmEyYTY3MDQ2MGEwZGE5ZGI4IiwidGFnIjoiIn0=; missav_session=eyJpdiI6IjBuWDlteG5vTlhmS2ppYk5oM0pSckE9PSIsInZhbHVlIjoiVGt4VjliYVJoc1RZRWZOQVBRSitnd09JOWFyVkY5TzBJMEt6RWVVeUs4cjlxcnkreFArUUNuU0JOMVFseXkwNDNzR0hEREN0emJyb2VSbWtvL3FIWjVmQWlDalJtRTM2QmhrTnNzenppQmQ5RDAxaHBtWkVGcWQ0SkNEZXhXUGoiLCJtYWMiOiJlYWQ0MGU2MGRmMjE4NWY2MGJiNDVjZDkzMDViY2JjYTYzZGI2OTIxOTI1OGQ2NDY5YjAwMmNmNzBhNThiM2NmIiwidGFnIjoiIn0=; CdHKodX58fXKDhwjz5xS1bqwjfrwa4oRgSDiOizw=eyJpdiI6IjMvSXlyb3RtcklSeUtxSEhnR0g5SGc9PSIsInZhbHVlIjoiQ3U5M3Q0UDVsWnFmQXRodnNqVTVNQW5rUGZwV3NCTlJheWJlS00rNXBIZnI5YzZmTFBTMUNRcWd1S1puaTk5a1RkckhXd0tNQi9IWmJJc3RWUVRGcjJFL1cvT0tLd0draExTTXlYZHdWa2dHaGw5dEJLTXI5WHFtNWVDTGErWlhUMllPN0FEL1Q2Q29kVUh1WHdlMzFxZ1dwVmcrSzlGSXA0aVB3ZGxzK0JoU09XTHU1N0ZJNXVRbHFHeWxSQTM0SlFMTWthTDBhampwdTVNUWp6YmttQ1p3cDFzdzRnVWY4UjVFRktJU2JOZ3BsK2RrZzJUayt2ZG5XZHF1cUVNbEd1UjBleHovamhqU3pid0VyQ1ZnUmFHYlJDNmxyS2t6V0xsQzVkY1WiTDRDSVg3MlArWXo4Qk4xWk5jQUpNQTQvWjdlTDBPMHhxSE1iMGxIeS9aTW1zRUlyU2RxVldaS0RqaFNUekM2TVpmMXM3Ry9IMHI3RjNsVHJwRU9QWm5mcGs5TzlKYkMrbkc1cTNkb3JacFNtZz09IiwibWFjIjoiZWRmMDVkMmQxYWJlNWE3ZDY4ZGYyMTk0MGYzN2YwMzUyMTQwNDFmNTI5YTY1MTM1OTc4ZWQwODdhMzg2ZDMyMSIsInRhZyI6IiJ9'
}

url = 'https://missav.ws/actresses/ranking'

try:
    r = requests.get(url, headers=headers)
    r.raise_for_status()
    html = r.text

    __name__ = re.findall('<h4 class="text-nord13 truncate">(.*?)</h4>', html)

    book = openpyxl.Workbook()
    sheet = book.active

    sheet['A1'] = '女优名字'
    sheet['B2'] = '片名'
    row = 3

    for i in range(len(__name__)):
        name = __name__[i]
        actress_url = f'https://missav.ws/actresses/{name}?sort=views&page=1'
        print(f"正在處理：{name} - {actress_url}")

        try:
            r_actress = requests.get(actress_url, headers=headers)
            r_actress.raise_for_status()
            html_actress = r_actress.text

            title = re.findall('<a class="text-secondary group-hover:text-primary" href=".*?">(.*?)</a>', html_actress, re.S)

            for a in range(len(title)):
                sheet.cell(row, 1).value = name
                sheet.cell(row, 2).value = title[a]
                row += 1

        except requests.exceptions.RequestException as e:
            print(f"女優 {name} 頁面抓取錯誤: {e}")
            continue

except requests.exceptions.RequestException as e:
    print(f"首頁抓取錯誤: {e}")
    exit()

except Exception as e:
    print(f"An error occurred: {e}")
    exit()

book.save('AV女优大全.xlsx')
print("資料抓取完成，儲存於 AV女优大全.xlsx")
